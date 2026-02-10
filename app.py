"""
MI Census Pro - Streamlit Cloud Edition
Version: V200_LIGHTWEIGHT
NO matplotlib dependency - pure Streamlit
"""

import streamlit as st
import pandas as pd
import numpy as np
import json
import os
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional
import warnings
warnings.filterwarnings('ignore')

# ========================================
# CONFIG
# ========================================
class AppConfig:
    VERSION = "V200_LIGHTWEIGHT"
    PASSWORD = "mandya"
    
    USERS = {
        "Chethan_NGM": "Nagamangala Taluk",
        "Gangadhar_MLV": "Malavalli Taluk",
        "Nagarjun_KRP": "K.R. Pete Taluk",
        "Prashanth_SRP": "Srirangapatna Taluk",
        "Purushottam_PDV": "Pandavapura Taluk",
        "Siddaraju_MDY": "Mandya Taluk",
        "Sunil_MDR": "Maddur Taluk",
        "Mandya_Admin": "District Admin"
    }
    
    COLORS = {
        "primary": "#1a73e8",
        "success": "#34A853",
        "warning": "#FBBC04",
        "danger": "#EA4335",
    }
    
    TALUKS = [
        "K.R. Pete Taluk",
        "Maddur Taluk",
        "Malavalli Taluk",
        "Mandya Taluk",
        "Nagamangala Taluk",
        "Pandavapura Taluk",
        "Srirangapatna Taluk"
    ]

# ========================================
# SETUP
# ========================================
def setup_streamlit_config():
    """Create optimal Streamlit config"""
    config_dir = ".streamlit"
    config_path = os.path.join(config_dir, "config.toml")
    
    if not os.path.exists(config_path):
        os.makedirs(config_dir, exist_ok=True)
        config = """
[server]
maxUploadSize = 100
maxMessageSize = 100
enableCORS = false
enableXsrfProtection = true
headless = true

[theme]
primaryColor = "#1a73e8"
backgroundColor = "#ffffff"
secondaryBackgroundColor = "#f8f9fa"

[client]
toolbarMode = "minimal"
showErrorDetails = false

[logger]
level = "error"

[cache]
maxMegabytes = 32
ttlSeconds = 30
"""
        with open(config_path, 'w') as f:
            f.write(config)

setup_streamlit_config()

st.set_page_config(
    page_title="MI Census Pro",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ========================================
# SESSION STATE
# ========================================
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.user = None

# ========================================
# UTILITIES
# ========================================
def save_taluk_data(taluk: str, data: Dict):
    """Save taluk metrics"""
    os.makedirs("central_data", exist_ok=True)
    filepath = f"central_data/{taluk.replace(' ', '_')}.json"
    
    safe_data = {}
    for k, v in data.items():
        if isinstance(v, (np.integer, np.int64)):
            safe_data[k] = int(v)
        elif isinstance(v, (np.floating, np.float64)):
            safe_data[k] = float(v)
        else:
            safe_data[k] = v
    
    safe_data['timestamp'] = datetime.now().isoformat()
    
    try:
        with open(filepath, 'w') as f:
            json.dump(safe_data, f)
    except Exception as e:
        st.warning(f"Could not save data: {e}")

def load_taluk_data(taluk: str) -> Dict:
    """Load taluk metrics"""
    filepath = f"central_data/{taluk.replace(' ', '_')}.json"
    
    if os.path.exists(filepath):
        try:
            with open(filepath, 'r') as f:
                return json.load(f)
        except:
            pass
    
    return {
        "taluk": taluk,
        "total_villages": 0,
        "completed": 0,
        "gw": 0,
        "sw": 0,
        "wb": 0,
        "submitted": 0
    }

@st.cache_data(ttl=60, max_entries=1)
def process_report(df_assign, df_monitor, taluk_name, completed, submitted):
    """Process report data - no image generation"""
    try:
        df_assign.columns = df_assign.columns.str.strip()
        df_monitor.columns = df_monitor.columns.str.strip()
        
        total = len(df_monitor)
        
        gw = pd.to_numeric(
            df_monitor.get('Total schedules GW', pd.Series([0]*total)),
            errors='coerce'
        ).fillna(0).sum()
        
        sw = pd.to_numeric(
            df_monitor.get('Total schedules SW', pd.Series([0]*total)),
            errors='coerce'
        ).fillna(0).sum()
        
        wb = pd.to_numeric(
            df_monitor.get('Total schedules WB', pd.Series([0]*total)),
            errors='coerce'
        ).fillna(0).sum()
        
        metrics = {
            "total_villages": int(total),
            "completed": int(completed),
            "submitted": int(submitted),
            "gw": int(gw),
            "sw": int(sw),
            "wb": int(wb)
        }
        
        return metrics, df_monitor.head(10)  # Return sample for display
        
    except Exception as e:
        st.error(f"Error processing data: {str(e)}")
        return None, None

def generate_excel_report(df_data: pd.DataFrame, metrics: Dict, taluk_name: str) -> bytes:
    """Generate Excel report"""
    try:
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': ['Total Villages', 'GW Schemes', 'SW Schemes', 'Wells', 'Completed', 'Submitted'],
                'Value': [
                    metrics['total_villages'],
                    metrics['gw'],
                    metrics['sw'],
                    metrics['wb'],
                    metrics['completed'],
                    metrics['submitted']
                ]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
            
            # Data sheet
            df_data.to_excel(writer, sheet_name='Data', index=False)
            
            # Format summary sheet
            ws = writer.sheets['Summary']
            header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            header_font = Font(bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
        
        output.seek(0)
        return output.getvalue()
        
    except Exception as e:
        st.error(f"Excel generation error: {str(e)}")
        return None

def create_status_card_html(taluk_name: str, metrics: Dict) -> str:
    """Create HTML status card"""
    total = metrics['total_villages']
    completed = metrics['completed']
    pct = (completed / total * 100) if total > 0 else 0
    
    html = f"""
    <div style="
        border: 3px solid {AppConfig.COLORS['primary']};
        border-radius: 10px;
        padding: 20px;
        background: #f8f9fa;
        margin: 10px 0;
    ">
        <h2 style="color: {AppConfig.COLORS['primary']}; margin: 0;">{taluk_name}</h2>
        <hr style="border: none; border-top: 2px solid {AppConfig.COLORS['primary']};">
        
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; margin: 20px 0;">
            <div style="text-align: center;">
                <p style="color: #666; font-size: 12px; margin: 0;">Total Villages</p>
                <p style="font-size: 32px; font-weight: bold; color: {AppConfig.COLORS['primary']}; margin: 10px 0;">{total}</p>
            </div>
            <div style="text-align: center;">
                <p style="color: #666; font-size: 12px; margin: 0;">Completed</p>
                <p style="font-size: 32px; font-weight: bold; color: {AppConfig.COLORS['success']}; margin: 10px 0;">{completed}</p>
            </div>
            <div style="text-align: center;">
                <p style="color: #666; font-size: 12px; margin: 0;">Completion %</p>
                <p style="font-size: 32px; font-weight: bold; color: {AppConfig.COLORS['warning']}; margin: 10px 0;">{pct:.1f}%</p>
            </div>
        </div>
        
        <div style="display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 15px; margin-top: 15px;">
            <div style="background: white; padding: 10px; border-radius: 5px; text-align: center;">
                <p style="color: #666; font-size: 11px; margin: 0;">GW Schemes</p>
                <p style="font-size: 20px; font-weight: bold; margin: 5px 0;">{metrics['gw']}</p>
            </div>
            <div style="background: white; padding: 10px; border-radius: 5px; text-align: center;">
                <p style="color: #666; font-size: 11px; margin: 0;">SW Schemes</p>
                <p style="font-size: 20px; font-weight: bold; margin: 5px 0;">{metrics['sw']}</p>
            </div>
            <div style="background: white; padding: 10px; border-radius: 5px; text-align: center;">
                <p style="color: #666; font-size: 11px; margin: 0;">Wells</p>
                <p style="font-size: 20px; font-weight: bold; margin: 5px 0;">{metrics['wb']}</p>
            </div>
        </div>
    </div>
    """
    return html

# ========================================
# MAIN APP
# ========================================
def main():
    st.markdown("# 🏛️ MI Census Pro")
    st.markdown("---")
    
    # Sidebar Login
    with st.sidebar:
        st.write("### 🔐 Login")
        
        col1, col2 = st.columns([1, 1])
        with col1:
            user = st.selectbox(
                "Officer:",
                list(AppConfig.USERS.keys()),
                label_visibility="collapsed"
            )
        
        with col2:
            password = st.text_input(
                "Password:",
                type="password",
                label_visibility="collapsed"
            )
        
        if st.button("Login", key="login_btn"):
            if password == AppConfig.PASSWORD:
                st.session_state.authenticated = True
                st.session_state.user = user
                st.success(f"✅ Welcome, {user}!")
                st.rerun()
            else:
                st.error("❌ Invalid password")
        
        if st.session_state.authenticated:
            st.write(f"**User:** {st.session_state.user}")
            if st.button("Logout"):
                st.session_state.authenticated = False
                st.rerun()
    
    # Check Authentication
    if not st.session_state.authenticated:
        st.warning("⚠️ Please login to continue")
        st.info("**Demo Credentials:**\n- User: Any officer\n- Password: mandya")
        return
    
    # Main Content
    tab1, tab2, tab3 = st.tabs(["📊 Report", "📈 Dashboard", "⚙️ Admin"])
    
    # TAB 1: REPORT
    with tab1:
        st.write("### 📊 Generate Report")
        
        col1, col2 = st.columns(2)
        with col1:
            taluk = st.selectbox("Select Taluk:", AppConfig.TALUKS)
        with col2:
            st.write("")
        
        assign_file = st.file_uploader("📁 Master File (Assignment):", type=["xlsx", "csv", "xls"], key="assign")
        monitor_file = st.file_uploader("📁 Monitor File (Data):", type=["xlsx", "csv", "xls"], key="monitor")
        
        col1, col2 = st.columns(2)
        with col1:
            completed = st.number_input("Completed Villages:", min_value=0, step=1, value=0)
        with col2:
            submitted = st.number_input("Submitted Villages:", min_value=0, step=1, value=0)
        
        if st.button("🚀 Generate Report", key="gen_report"):
            if assign_file and monitor_file:
                with st.spinner("⏳ Processing..."):
                    try:
                        # Load files
                        if assign_file.name.endswith(('.xlsx', '.xls')):
                            df_assign = pd.read_excel(assign_file)
                        else:
                            df_assign = pd.read_csv(assign_file)
                        
                        if monitor_file.name.endswith(('.xlsx', '.xls')):
                            df_monitor = pd.read_excel(monitor_file)
                        else:
                            df_monitor = pd.read_csv(monitor_file)
                        
                        # Process
                        metrics, df_data = process_report(df_assign, df_monitor, taluk, int(completed), int(submitted))
                        
                        if metrics:
                            # Save data
                            save_taluk_data(taluk, metrics)
                            
                            # Display results
                            st.success("✅ Report generated successfully!")
                            
                            # Metrics
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Total Villages", metrics['total_villages'])
                            with col2:
                                st.metric("GW Schemes", metrics['gw'])
                            with col3:
                                st.metric("SW Schemes", metrics['sw'])
                            with col4:
                                st.metric("Wells", metrics['wb'])
                            
                            st.write("")
                            
                            # Status card
                            st.markdown(create_status_card_html(taluk, metrics), unsafe_allow_html=True)
                            
                            st.write("")
                            
                            # Data preview
                            if df_data is not None and len(df_data) > 0:
                                st.write("#### 📋 Data Preview")
                                st.dataframe(df_data.head(), use_container_width=True)
                            
                            st.write("")
                            
                            # Excel download
                            excel_bytes = generate_excel_report(df_data, metrics, taluk)
                            if excel_bytes:
                                st.download_button(
                                    "📥 Download Excel Report",
                                    excel_bytes,
                                    f"{taluk}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                    
                    except Exception as e:
                        st.error(f"❌ Error: {str(e)}")
                        st.info("Make sure both files are in correct Excel or CSV format")
            else:
                st.warning("⚠️ Please upload both files")
    
    # TAB 2: DASHBOARD
    with tab2:
        st.write("### 📈 District Dashboard")
        
        # Load all taluk data
        all_data = []
        for taluk in AppConfig.TALUKS:
            data = load_taluk_data(taluk)
            all_data.append(data)
        
        if all_data:
            df_dashboard = pd.DataFrame(all_data)
            
            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Taluks", len(df_dashboard))
            with col2:
                total_villages = int(df_dashboard['total_villages'].sum())
                st.metric("Total Villages", total_villages)
            with col3:
                total_completed = int(df_dashboard['completed'].sum())
                st.metric("Total Completed", total_completed)
            with col4:
                if total_villages > 0:
                    pct = (total_completed / total_villages * 100)
                    st.metric("Overall %", f"{pct:.1f}%")
                else:
                    st.metric("Overall %", "0%")
            
            st.write("")
            
            # Table
            st.write("#### Taluk-wise Summary")
            display_cols = ['taluk', 'total_villages', 'completed', 'gw', 'sw', 'wb', 'submitted']
            st.dataframe(
                df_dashboard[display_cols],
                use_container_width=True,
                hide_index=True
            )
            
            st.write("")
            
            # Progress bars
            st.write("#### Completion Progress")
            for idx, row in df_dashboard.iterrows():
                if row['total_villages'] > 0:
                    progress = row['completed'] / row['total_villages']
                    st.progress(
                        progress,
                        text=f"{row['taluk']}: {row['completed']}/{row['total_villages']} ({progress*100:.1f}%)"
                    )
        else:
            st.info("ℹ️ No data available. Generate reports to see dashboard.")
    
    # TAB 3: ADMIN
    with tab3:
        st.write("### ⚙️ Admin Functions")
        
        if st.session_state.user == "Mandya_Admin":
            st.write("#### Administrator Panel")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                if st.button("🔄 Clear Cache"):
                    st.cache_data.clear()
                    st.success("✅ Cache cleared")
            
            with col2:
                if st.button("📥 Export All Data"):
                    all_data = []
                    for taluk in AppConfig.TALUKS:
                        data = load_taluk_data(taluk)
                        all_data.append(data)
                    
                    df_export = pd.DataFrame(all_data)
                    csv = df_export.to_csv(index=False)
                    
                    st.download_button(
                        "📊 Download CSV",
                        csv,
                        f"census_export_{datetime.now().strftime('%Y%m%d')}.csv",
                        "text/csv"
                    )
            
            with col3:
                if st.button("🗑️ Reset All Data"):
                    if st.checkbox("Confirm reset all data"):
                        import shutil
                        try:
                            shutil.rmtree("central_data")
                            st.success("✅ All data reset")
                        except:
                            st.info("No data to reset")
            
            st.write("")
            st.divider()
            st.write("")
            
            st.info(f"**Version:** {AppConfig.VERSION}")
            st.info(f"**Server Time:** {datetime.now(timezone.utc) + timedelta(hours=5.5)}")
            st.info(f"**Total Officers:** {len(AppConfig.USERS)}")
            
        else:
            st.warning("⚠️ Admin access only")
            st.info(f"Current user: {st.session_state.user}")

if __name__ == "__main__":
    main()
